"""Sales report upload: normalize, persist, and derive analytics preview."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from graphql.data_sources import AnalyticsRun, Location, SessionLocal
from graphql.reports import (
    Order,
    line_items_to_orders,
    normalize_sales_report,
    persist_sales_report,
    run_sales_analytics,
)
from graphql.reports.ingest import NormalizedLineItemData
from graphql.schema.auth import require_location_owner


@dataclass
class SalesReportIngestResult:
    """Domain result mapped to GraphQL `ExcelUploadResult` in the resolver."""

    filename: str
    sheet_names: list[str]
    header_preview: list[str]
    size_bytes: int
    normalized_rows: list[NormalizedLineItemData]
    orders: list[Order]
    sales_analytics: dict[str, Any]


def ingest_sales_report_upload(
    *,
    payload: bytes,
    filename: str | None,
    location_id: int,
    user_id: str,
) -> SalesReportIngestResult:
    """Create analytics run, persist order facts in one DB session, then build preview data."""
    normalized_rows_data, detected_pos = normalize_sales_report(payload)
    with SessionLocal() as session:
        period_start: datetime | None = None
        period_end: datetime | None = None
        if normalized_rows_data:
            times = [
                row.orderTime
                if isinstance(row.orderTime, datetime)
                else datetime.fromisoformat(str(row.orderTime))
                for row in normalized_rows_data
            ]
            period_start = min(times)
            period_end = max(times)

        loc = session.get(Location, int(location_id))
        if loc is None:
            raise ValueError(f"Location {location_id} not found")
        require_location_owner(session, int(location_id), user_id)

        analytics_run = AnalyticsRun(
            name=filename or "sales_report",
            filename=filename or "",
            pos_system=detected_pos,
            period_start=period_start.date() if period_start else None,
            period_end=period_end.date() if period_end else None,
            location_id=loc.id,
        )
        session.add(analytics_run)
        session.commit()
        session.refresh(analytics_run)

        persist_sales_report(
            session,
            normalized_rows_data,
            detected_pos,
            analytics_run_id=analytics_run.id,
        )

    orders_data = line_items_to_orders(normalized_rows_data)
    sales_analytics_dict: dict[str, Any] = run_sales_analytics(normalized_rows_data)

    workbook = load_workbook(filename=BytesIO(payload), read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    header_preview: list[str] = []

    if sheet_names:
        active_sheet = workbook[sheet_names[0]]
        first_row = next(active_sheet.iter_rows(max_row=1, values_only=True), ())
        header_preview = [str(value) if value is not None else "" for value in first_row]

    workbook.close()

    return SalesReportIngestResult(
        filename=filename or "",
        sheet_names=sheet_names,
        header_preview=header_preview,
        size_bytes=len(payload),
        normalized_rows=normalized_rows_data,
        orders=orders_data,
        sales_analytics=sales_analytics_dict,
    )
