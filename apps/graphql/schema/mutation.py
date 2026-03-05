from dataclasses import asdict
from datetime import datetime
from io import BytesIO
from typing import Any

import strawberry
from openpyxl import load_workbook
from strawberry.file_uploads import Upload
from strawberry.scalars import JSON

from graphql.reports import (
    Order,
    line_items_to_orders,
    normalize_sales_report,
    persist_sales_report,
    run_sales_analytics,
)


@strawberry.type
class NormalizedLineItem:
    billNumber: str
    menu: str
    qty: int
    price: float
    totalAfterBillDiscount: float
    orderTime: datetime
    menuCategory: str
    menuCategoryDetail: str


@strawberry.type
class OrderItemType:
    menu: str
    qty: int
    price: float
    totalAfterBillDiscount: float
    menuCategory: str
    menuCategoryDetail: str


@strawberry.type
class OrderType:
    billNumber: str
    orderTime: datetime
    items: list[OrderItemType]


@strawberry.type
class ExcelUploadResult:
    filename: str
    sheet_names: list[str]
    header_preview: list[str]
    size_bytes: int
    normalized_rows: list[NormalizedLineItem]
    orders: list[OrderType]
    sales_analytics: JSON


def _order_to_strawberry(order: Order) -> OrderType:
    return OrderType(
        billNumber=order.billNumber,
        orderTime=order.orderTime,
        items=[
            OrderItemType(
                menu=item.menu,
                qty=item.qty,
                price=item.price,
                totalAfterBillDiscount=item.totalAfterBillDiscount,
                menuCategory=item.menuCategory,
                menuCategoryDetail=item.menuCategoryDetail,
            )
            for item in order.items
        ],
    )


@strawberry.type
class Mutation:
    @strawberry.mutation
    async def upload_sales_report(self, file: Upload) -> ExcelUploadResult:
        payload = await file.read()
        normalized_rows_data, detected_pos = normalize_sales_report(payload)
        persist_sales_report(normalized_rows_data, detected_pos)

        normalized_rows = [
            NormalizedLineItem(**asdict(row)) for row in normalized_rows_data
        ]

        orders_data = line_items_to_orders(normalized_rows_data)
        orders = [_order_to_strawberry(o) for o in orders_data]

        sales_analytics_dict: dict[str, Any] = run_sales_analytics(normalized_rows_data)

        workbook = load_workbook(
            filename=BytesIO(payload), read_only=True, data_only=True
        )
        sheet_names = workbook.sheetnames
        header_preview: list[str] = []

        if sheet_names:
            active_sheet = workbook[sheet_names[0]]
            first_row = next(active_sheet.iter_rows(max_row=1, values_only=True), ())
            header_preview = [
                str(value) if value is not None else "" for value in first_row
            ]

        workbook.close()

        return ExcelUploadResult(
            filename=file.filename,
            sheet_names=sheet_names,
            header_preview=header_preview,
            size_bytes=len(payload),
            normalized_rows=normalized_rows,
            orders=orders,
            sales_analytics=sales_analytics_dict,
        )
